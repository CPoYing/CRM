import io
import json
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# Page config
# ============================================================
APP_NAME = "專案管理開發時程"

st.set_page_config(
    page_title=APP_NAME,
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# Constants
# ============================================================
COLUMNS = [
    "WBS編號", "層級", "任務名稱", "負責人",
    "開始日期", "截止日期", "狀態", "優先順序", "進度(%)", "備註",
]
DATE_COLS = ["開始日期", "截止日期"]
TEXT_COLS = ["WBS編號", "層級", "任務名稱", "負責人", "狀態", "優先順序", "備註"]

STATUS_OPTIONS = ["未開始", "進行中", "已完成", "暫停", "延遲"]
PRIORITY_OPTIONS = ["高", "中", "低"]
LEVEL_OPTIONS = ["主任務", "子任務"]

STATUS_COLORS = {
    "已完成": "#70AD47",
    "進行中": "#5B9BD5",
    "延遲": "#ED7D31",
    "未開始": "#A6A6A6",
    "暫停": "#FF6B6B",
}

NEW_ROW_DEFAULTS = {
    "WBS編號": "", "層級": "子任務", "任務名稱": "", "負責人": "",
    "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中",
    "進度(%)": 0, "備註": "",
}

DATA_FILE = Path(__file__).parent / "data" / "tasks.json"
DEFAULT_PROJECT_NAME = "訂單管理系統模組開發"

# ============================================================
# Default data
# ============================================================
DEFAULT_DATA = [
    {"WBS編號": "1", "層級": "主任務", "任務名稱": "合約", "開始日期": "2026-05-01", "截止日期": "2026-06-30", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "1.1", "層級": "子任務", "任務名稱": "合約建立", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "1.2", "層級": "子任務", "任務名稱": "合約修改", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "1.3", "層級": "子任務", "任務名稱": "保證票(存入/取出)", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "2", "層級": "主任務", "任務名稱": "訂單", "開始日期": "2026-06-01", "截止日期": "2026-07-31", "狀態": "延遲", "優先順序": "高", "進度(%)": 80, "備註": ""},
    {"WBS編號": "2.1", "層級": "子任務", "任務名稱": "訂單建立", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "2.2", "層級": "子任務", "任務名稱": "訂單修改", "開始日期": "", "截止日期": "", "狀態": "延遲", "優先順序": "高", "進度(%)": 60, "備註": ""},
    {"WBS編號": "2.3", "層級": "子任務", "任務名稱": "訂單承約參數設定", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": ""},
    {"WBS編號": "2.4", "層級": "子任務", "任務名稱": "有主調撥單", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "高", "進度(%)": 100, "備註": "很久以前有試過OK，但時間久遠可能要在試一次"},
    {"WBS編號": "3", "層級": "主任務", "任務名稱": "出貨", "開始日期": "2026-07-01", "截止日期": "2026-08-31", "狀態": "進行中", "優先順序": "高", "進度(%)": 70, "備註": ""},
    {"WBS編號": "3.1", "層級": "子任務", "任務名稱": "出貨申請", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 70, "備註": "SS YM可以 SC不行/SW 兩廠都不行"},
    {"WBS編號": "3.2", "層級": "子任務", "任務名稱": "出貨修改", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 70, "備註": ""},
    {"WBS編號": "3.3", "層級": "子任務", "任務名稱": "信用狀", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "低", "進度(%)": 70, "備註": ""},
    {"WBS編號": "4", "層級": "主任務", "任務名稱": "點銅", "開始日期": "2026-08-01", "截止日期": "2026-08-31", "狀態": "進行中", "優先順序": "高", "進度(%)": 0, "備註": ""},
    {"WBS編號": "4.1", "層級": "子任務", "任務名稱": "點銅單建立", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 0, "備註": ""},
    {"WBS編號": "4.2", "層級": "子任務", "任務名稱": "點銅單修改", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 0, "備註": ""},
    {"WBS編號": "5", "層級": "主任務", "任務名稱": "出廠文件", "開始日期": "2026-06-01", "截止日期": "2026-07-31", "狀態": "延遲", "優先順序": "高", "進度(%)": 50, "備註": ""},
    {"WBS編號": "5.1", "層級": "子任務", "任務名稱": "出廠文件申請", "開始日期": "", "截止日期": "", "狀態": "延遲", "優先順序": "高", "進度(%)": 50, "備註": ""},
    {"WBS編號": "6", "層級": "主任務", "任務名稱": "送審文件", "開始日期": "2026-06-01", "截止日期": "2026-06-30", "狀態": "已完成", "優先順序": "中", "進度(%)": 100, "備註": ""},
    {"WBS編號": "6.1", "層級": "子任務", "任務名稱": "送審文件申請", "開始日期": "", "截止日期": "", "狀態": "已完成", "優先順序": "中", "進度(%)": 100, "備註": ""},
    {"WBS編號": "7", "層級": "主任務", "任務名稱": "信用額度", "開始日期": "2026-07-01", "截止日期": "2026-09-30", "狀態": "進行中", "優先順序": "中", "進度(%)": 60, "備註": ""},
    {"WBS編號": "7.1", "層級": "子任務", "任務名稱": "信用額度申請", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "中", "進度(%)": 60, "備註": ""},
    {"WBS編號": "7.2", "層級": "子任務", "任務名稱": "信用額度-其他申請", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "中", "進度(%)": 60, "備註": ""},
    {"WBS編號": "8", "層級": "主任務", "任務名稱": "強制放行", "開始日期": "2026-07-01", "截止日期": "2026-08-31", "狀態": "進行中", "優先順序": "高", "進度(%)": 60, "備註": ""},
    {"WBS編號": "8.1", "層級": "子任務", "任務名稱": "強制放行申請", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 60, "備註": ""},
    {"WBS編號": "8.2", "層級": "子任務", "任務名稱": "強制放行-審核", "開始日期": "", "截止日期": "", "狀態": "進行中", "優先順序": "高", "進度(%)": 60, "備註": ""},
    {"WBS編號": "9", "層級": "主任務", "任務名稱": "一般寄庫裁線", "開始日期": "2026-09-01", "截止日期": "2026-09-30", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "9.1", "層級": "子任務", "任務名稱": "一般寄庫裁線-申請", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "9.2", "層級": "子任務", "任務名稱": "一般寄庫裁線-出貨申請", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "9.3", "層級": "子任務", "任務名稱": "一般寄庫品查詢", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10", "層級": "主任務", "任務名稱": "大軸寄庫裁線", "開始日期": "2026-09-20", "截止日期": "2026-10-31", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.1", "層級": "子任務", "任務名稱": "大軸品下單", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.2", "層級": "子任務", "任務名稱": "大軸品轉寄庫品", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.3", "層級": "子任務", "任務名稱": "大軸寄庫裁線", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.4", "層級": "子任務", "任務名稱": "大軸寄庫裁線出貨", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.5", "層級": "子任務", "任務名稱": "大軸寄庫品查詢", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
    {"WBS編號": "10.6", "層級": "子任務", "任務名稱": "大軸寄庫品明細", "開始日期": "", "截止日期": "", "狀態": "未開始", "優先順序": "中", "進度(%)": 0, "備註": ""},
]


# ============================================================
# Data helpers
# ============================================================
def normalize_df(df):
    """把任意來源的資料整理成固定欄位與型別，避免下游元件炸掉。"""
    df = df.copy()
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNS]

    for col in DATE_COLS:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    # 容忍上傳檔裡寫成 "60%" 或 " 60 " 的進度
    progress = df["進度(%)"].astype(str).str.replace("%", "", regex=False).str.strip()
    df["進度(%)"] = (
        pd.to_numeric(progress, errors="coerce").fillna(0).clip(0, 100).astype(int)
    )

    for col in TEXT_COLS:
        df[col] = df[col].fillna("").astype(str).str.strip()

    # SelectboxColumn 遇到不在選項內的值會整格報錯，先收斂掉
    df["層級"] = df["層級"].where(df["層級"].isin(LEVEL_OPTIONS), "子任務")
    df["狀態"] = df["狀態"].where(df["狀態"].isin(STATUS_OPTIONS), "未開始")
    df["優先順序"] = df["優先順序"].where(df["優先順序"].isin(PRIORITY_OPTIONS), "中")

    return df.reset_index(drop=True)


def df_to_records(df):
    """轉成可 JSON 序列化的 list[dict]（日期輸出 YYYY-MM-DD 或空字串）。"""
    records = []
    for _, row in df.iterrows():
        rec = {}
        for col in COLUMNS:
            val = row[col]
            if col in DATE_COLS:
                rec[col] = "" if pd.isna(val) else pd.Timestamp(val).strftime("%Y-%m-%d")
            elif col == "進度(%)":
                rec[col] = int(val)
            else:
                rec[col] = "" if pd.isna(val) else str(val)
        records.append(rec)
    return records


def save_data():
    """把目前資料寫到本機 data/tasks.json，重新整理瀏覽器也不會不見。"""
    try:
        DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "project_name": st.session_state.project_name,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "tasks": df_to_records(st.session_state.df),
        }
        DATA_FILE.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        st.session_state.last_saved = payload["saved_at"]
        return True
    except Exception as exc:  # 存檔失敗不該讓整個 App 掛掉
        st.session_state.save_error = str(exc)
        return False


def load_data():
    """讀本機存檔；沒有存檔就用內建預設資料。"""
    if DATA_FILE.exists():
        try:
            payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            df = normalize_df(pd.DataFrame(payload.get("tasks", [])))
            if df.empty:
                raise ValueError("存檔內沒有任務")
            return df, payload.get("project_name", DEFAULT_PROJECT_NAME)
        except Exception as exc:
            st.warning(f"讀取本機存檔失敗，改用預設資料：{exc}")
    return normalize_df(pd.DataFrame(DEFAULT_DATA)), DEFAULT_PROJECT_NAME


def init_session():
    if "df" not in st.session_state:
        df, project_name = load_data()
        st.session_state.df = df
        st.session_state.project_name = project_name
    st.session_state.setdefault("editor_ver", 0)
    st.session_state.setdefault("last_view_index", None)
    st.session_state.setdefault("uploaded_token", None)
    st.session_state.setdefault("last_saved", None)
    st.session_state.setdefault("save_error", None)
    st.session_state.setdefault("collapsed_wbs", set())   # 被收合的主任務 WBS
    st.session_state.setdefault("expand_ver", 0)          # 收合控制項的 widget 版本
    st.session_state.setdefault("expand_options", [])


def editor_key():
    return f"task_editor_{st.session_state.editor_ver}"


def wbs_sort_key(value):
    """自然排序：1 < 1.1 < 2 < 10（純字串排序會把 10 排到 2 前面）。"""
    parts = []
    for chunk in str(value).split("."):
        chunk = chunk.strip()
        parts.append(int(chunk) if chunk.isdigit() else 10**9)
    return parts


def parent_wbs(value):
    text = str(value).strip()
    return text.rsplit(".", 1)[0] if "." in text else None


def child_counts(df):
    """每個主任務底下的子任務數量 {WBS: n}。"""
    counts = {}
    sub_wbs = df.loc[df["層級"] == "子任務", "WBS編號"].astype(str).str.strip()
    for wbs in df.loc[df["層級"] == "主任務", "WBS編號"].astype(str).str.strip():
        counts[wbs] = int(sub_wbs.str.startswith(wbs + ".").sum())
    return counts


def collapsible_mains(df):
    """可收合的主任務（有子任務的才需要）→ [(wbs, 顯示文字, 子任務數)]。"""
    counts = child_counts(df)
    mains = []
    for _, row in df[df["層級"] == "主任務"].iterrows():
        wbs = str(row["WBS編號"]).strip()
        if counts.get(wbs):
            mains.append((wbs, f"{wbs} {row['任務名稱']}（{counts[wbs]}）", counts[wbs]))
    mains.sort(key=lambda item: wbs_sort_key(item[0]))
    return mains


def collapse_mask(df, collapsed):
    """被收合的主任務，其子任務不顯示。"""
    if not collapsed:
        return pd.Series(True, index=df.index)
    wbs = df["WBS編號"].astype(str).str.strip()
    hidden = pd.Series(False, index=df.index)
    for parent in collapsed:
        hidden |= (df["層級"] == "子任務") & wbs.str.startswith(parent + ".")
    return ~hidden


def orphan_subtasks(df):
    """找不到對應主任務的子任務（例如主任務被刪掉了）。這些列不會被收合，永遠看得到。"""
    main_wbs = set(df.loc[df["層級"] == "主任務", "WBS編號"].astype(str).str.strip())
    subs = df[df["層級"] == "子任務"]
    if subs.empty:
        return subs
    return subs[~subs["WBS編號"].map(parent_wbs).isin(main_wbs)]


def tree_markers(view_df, collapsed, counts):
    """任務清單最左邊的樹狀符號。"""
    marks = []
    for _, row in view_df.iterrows():
        if row["層級"] != "主任務":
            marks.append("　└")
            continue
        wbs = str(row["WBS編號"]).strip()
        n = counts.get(wbs, 0)
        if not n:
            marks.append("")
        elif wbs in collapsed:
            marks.append(f"▶ {n}")
        else:
            marks.append("▼")
    return marks


def _same(a, b):
    if pd.isna(a) and pd.isna(b):
        return True
    if pd.isna(a) or pd.isna(b):
        return False
    return a == b


def apply_editor_changes():
    """
    把 data_editor 的編輯結果寫回主資料。

    直接讀 widget state（edited_rows / added_rows / deleted_rows），
    位置索引透過上一輪渲染時保存的 last_view_index 對應回主資料，
    所以「有篩選時刪除列」也會真的刪掉，不會只是畫面上消失。
    """
    view_index = st.session_state.last_view_index
    state = st.session_state.get(editor_key())
    if not state or view_index is None:
        return

    edited_rows = state.get("edited_rows") or {}
    added_rows = state.get("added_rows") or []
    deleted_rows = state.get("deleted_rows") or []

    df = st.session_state.df
    changed = False

    # 1) 既有列的欄位修改
    for pos, changes in edited_rows.items():
        pos = int(pos)
        if pos >= len(view_index):
            continue
        idx = view_index[pos]
        if idx not in df.index:
            continue
        for col, val in changes.items():
            if col not in COLUMNS:
                continue
            if col in DATE_COLS:
                val = pd.to_datetime(val, errors="coerce")
            elif col == "進度(%)":
                val = int(pd.to_numeric(val, errors="coerce") or 0)
            if not _same(df.at[idx, col], val):
                df.at[idx, col] = val
                changed = True

    # 2) 刪除列
    drop_idx = [
        view_index[int(pos)] for pos in deleted_rows if int(pos) < len(view_index)
    ]
    drop_idx = [i for i in drop_idx if i in df.index]
    if drop_idx:
        df = df.drop(index=drop_idx)
        changed = True

    # 3) 新增列（整列全空的先不寫入，讓使用者慢慢填）
    new_rows = []
    for raw in added_rows:
        raw = {k: v for k, v in (raw or {}).items() if k in COLUMNS}
        if not any(str(v).strip() for v in raw.values() if v is not None):
            continue
        rec = dict(NEW_ROW_DEFAULTS)
        rec.update(raw)
        new_rows.append(rec)
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
        changed = True

    if not changed:
        return

    st.session_state.df = normalize_df(df)
    save_data()

    # 有增刪列、或目前有套用篩選時，位置索引已經對不上，
    # 必須換掉 editor 的 key 清掉舊的編輯狀態，否則會套用到錯的列。
    if new_rows or drop_idx or st.session_state.get("filter_active"):
        st.session_state.editor_ver += 1
        st.session_state.last_view_index = None
        st.rerun()


# ============================================================
# Metrics / charts
# ============================================================
def get_main_tasks(df):
    return df[df["層級"] == "主任務"].copy()


def overdue_mask(df, today=None):
    today = pd.Timestamp(today or date.today())
    return df["截止日期"].notna() & (df["截止日期"] < today) & (df["狀態"] != "已完成")


def calc_metrics(df):
    main = get_main_tasks(df)
    total = len(main)
    done = int((main["狀態"] == "已完成").sum())
    return {
        "主任務數": total,
        "已完成": done,
        "進行中": int((main["狀態"] == "進行中").sum()),
        "延遲": int((main["狀態"] == "延遲").sum()),
        "未開始": int((main["狀態"] == "未開始").sum()),
        "完成率": (done / total * 100) if total else 0,
        "整體進度": main["進度(%)"].mean() if total else 0,
        "逾期": int(overdue_mask(df).sum()),
        "子任務數": int((df["層級"] == "子任務").sum()),
    }


def rollup_main_progress(df):
    """由子任務平均值重算主任務進度；沒有子任務的主任務維持原值。"""
    df = df.copy()
    for idx, row in df[df["層級"] == "主任務"].iterrows():
        wbs = str(row["WBS編號"]).strip()
        children = df[
            (df["層級"] == "子任務")
            & (df["WBS編號"].astype(str).str.strip().str.startswith(wbs + "."))
        ]
        if len(children):
            df.at[idx, "進度(%)"] = int(round(children["進度(%)"].mean()))
    return df


def build_gantt(df, show_subtasks=False, inherit_parent_dates=True):
    """Plotly timeline。只畫同時有開始與截止日期的任務。"""
    plot_df = df.copy()

    if inherit_parent_dates:
        parents = plot_df.set_index(plot_df["WBS編號"].astype(str).str.strip())
        for idx, row in plot_df.iterrows():
            if row["層級"] != "子任務":
                continue
            if pd.notna(row["開始日期"]) and pd.notna(row["截止日期"]):
                continue
            pw = parent_wbs(row["WBS編號"])
            if pw and pw in parents.index:
                parent = parents.loc[pw]
                if isinstance(parent, pd.DataFrame):
                    parent = parent.iloc[0]
                for col in DATE_COLS:
                    if pd.isna(plot_df.at[idx, col]):
                        plot_df.at[idx, col] = parent[col]

    plot_df = plot_df[plot_df["開始日期"].notna() & plot_df["截止日期"].notna()].copy()
    if not show_subtasks:
        plot_df = plot_df[plot_df["層級"] == "主任務"]
    if plot_df.empty:
        return None

    plot_df = plot_df.sort_values("WBS編號", key=lambda s: s.map(wbs_sort_key))
    plot_df["標籤"] = plot_df.apply(
        lambda r: f"{'　　' if r['層級'] == '子任務' else ''}{r['WBS編號']} {r['任務名稱']}",
        axis=1,
    )

    fig = px.timeline(
        plot_df,
        x_start="開始日期",
        x_end="截止日期",
        y="標籤",
        color="狀態",
        color_discrete_map=STATUS_COLORS,
        hover_data={
            "WBS編號": True,
            "層級": True,
            "負責人": True,
            "進度(%)": True,
            "優先順序": True,
            "標籤": False,
        },
    )

    # y 軸第一個分類預設在最下面，反轉陣列讓 WBS 小的排在最上面
    fig.update_yaxes(
        categoryorder="array",
        categoryarray=list(reversed(plot_df["標籤"].tolist())),
    )
    fig.update_layout(
        height=max(400, len(plot_df) * 28 + 120),
        margin=dict(l=20, r=20, t=30, b=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis_title=None,
        yaxis_title=None,
        font=dict(size=12),
    )
    fig.update_traces(marker_line_width=0, opacity=0.9)

    today = pd.Timestamp(date.today())
    if plot_df["開始日期"].min() <= today <= plot_df["截止日期"].max():
        fig.add_vline(x=today, line_width=2, line_dash="dash", line_color="#C00000")

    return fig


# ============================================================
# Excel export
# ============================================================
def to_excel_bytes(df, project_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "開發時程"

    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", size=9, bold=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    dark_blue = PatternFill("solid", fgColor="2F5496")
    light_blue = PatternFill("solid", fgColor="D6E3F0")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="C6EFCE")
    main_fill = PatternFill("solid", fgColor="DEEBF7")
    alt_fill = PatternFill("solid", fgColor="F8F9FA")

    last_col = get_column_letter(len(COLUMNS))

    # 標題列
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{APP_NAME} — {project_name}"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_blue
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 資訊列
    ws["A3"] = "專案名稱"
    ws["A3"].font = Font(name="Arial", size=10, bold=True)
    ws["A3"].fill = light_blue
    ws.merge_cells("B3:C3")
    ws["B3"] = project_name
    ws["B3"].fill = yellow
    ws["B3"].border = thin

    ws["D3"] = "匯出時間"
    ws["D3"].font = Font(name="Arial", size=10, bold=True)
    ws["D3"].fill = light_blue
    ws["E3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["E3"].fill = yellow
    ws["E3"].border = thin

    main = df[df["層級"] == "主任務"]
    ws["F3"] = "整體進度"
    ws["F3"].font = Font(name="Arial", size=10, bold=True)
    ws["F3"].fill = light_blue
    ws["G3"] = (main["進度(%)"].mean() if len(main) else 0) / 100
    ws["G3"].number_format = "0.0%"
    ws["G3"].fill = green
    ws["G3"].border = thin
    ws["G3"].font = Font(name="Arial", size=12, bold=True)

    # 表頭
    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=5, column=c, value=col)
        cell.font = header_font
        cell.fill = dark_blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[5].height = 24

    # 資料列：用 enumerate 產生列號，不依賴 DataFrame index
    for offset, (_, row) in enumerate(df.iterrows()):
        r = 6 + offset
        is_main = row["層級"] == "主任務"
        for c, col in enumerate(COLUMNS, 1):
            val = row[col]
            if col in DATE_COLS:
                val = None if pd.isna(val) else pd.Timestamp(val).date()
            elif col == "進度(%)":
                val = float(val) / 100
            elif val == "" or pd.isna(val):
                val = None

            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold_font if (is_main and col in ("WBS編號", "任務名稱")) else normal_font
            cell.border = thin
            if col == "進度(%)":
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")
            elif col in DATE_COLS:
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = Alignment(horizontal="center")

            if is_main:
                cell.fill = main_fill
            elif offset % 2 == 0:
                cell.fill = alt_fill

    widths = {
        "WBS編號": 10, "層級": 9, "任務名稱": 22, "負責人": 10,
        "開始日期": 12, "截止日期": 12, "狀態": 9, "優先順序": 9,
        "進度(%)": 9, "備註": 40,
    }
    for c, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 12)

    ws.freeze_panes = "A6"

    # 下拉驗證（欄位位置由 COLUMNS 推算，加欄位不會跑掉）
    last_row = 5 + max(len(df), 1)
    for col, options in (
        ("層級", LEVEL_OPTIONS),
        ("狀態", STATUS_OPTIONS),
        ("優先順序", PRIORITY_OPTIONS),
    ):
        letter = get_column_letter(COLUMNS.index(col) + 1)
        dv = DataValidation(
            type="list", formula1=f'"{",".join(options)}"', allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}6:{letter}{last_row}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# Main App
# ============================================================
def main():
    init_session()
    apply_editor_changes()  # 先套用上一輪的編輯，統計與甘特圖才會是最新的

    # ----- Sidebar -----
    with st.sidebar:
        st.title(f"📋 {APP_NAME}")
        st.caption("主任務 / 子任務可收合 + 甘特圖")

        name = st.text_input("專案名稱", value=st.session_state.project_name)
        if name != st.session_state.project_name:
            st.session_state.project_name = name
            save_data()

        st.divider()
        st.subheader("篩選")
        filter_level = st.multiselect("層級", LEVEL_OPTIONS, default=LEVEL_OPTIONS)
        filter_status = st.multiselect("狀態", STATUS_OPTIONS, default=STATUS_OPTIONS)
        filter_priority = st.multiselect("優先順序", PRIORITY_OPTIONS, default=PRIORITY_OPTIONS)

        st.divider()
        st.subheader("甘特圖選項")
        show_sub_in_gantt = st.checkbox("顯示子任務", value=False)
        inherit_dates = st.checkbox(
            "子任務沒填日期時沿用主任務日期", value=True,
            help="關掉的話，沒有日期的子任務就不會出現在甘特圖上",
        )

        st.divider()
        st.subheader("資料操作")

        uploaded = st.file_uploader("上傳 Excel / CSV 覆蓋資料", type=["xlsx", "csv"])
        if uploaded is not None:
            token = getattr(uploaded, "file_id", None) or f"{uploaded.name}:{uploaded.size}"
            # 同一個檔案只處理一次，否則每次互動都會把你的編輯蓋掉
            if token != st.session_state.uploaded_token:
                try:
                    uploaded.seek(0)
                    if uploaded.name.lower().endswith(".csv"):
                        new_df = pd.read_csv(uploaded)
                    else:
                        new_df = pd.read_excel(uploaded)
                    st.session_state.df = normalize_df(new_df)
                    st.session_state.uploaded_token = token
                    st.session_state.editor_ver += 1
                    st.session_state.last_view_index = None
                    save_data()
                    st.rerun()
                except Exception as exc:
                    st.session_state.uploaded_token = token
                    st.error(f"讀取失敗：{exc}")

        col_l, col_r = st.columns(2)
        if col_l.button("由子任務重算主任務進度", width="stretch"):
            st.session_state.df = rollup_main_progress(st.session_state.df)
            st.session_state.editor_ver += 1
            st.session_state.last_view_index = None
            save_data()
            st.rerun()

        if col_r.button("逾期標記為延遲", width="stretch"):
            mask = overdue_mask(st.session_state.df)
            st.session_state.df.loc[mask, "狀態"] = "延遲"
            st.session_state.editor_ver += 1
            st.session_state.last_view_index = None
            save_data()
            st.rerun()

        if st.button("還原預設資料", type="secondary", width="stretch"):
            st.session_state.df = normalize_df(pd.DataFrame(DEFAULT_DATA))
            st.session_state.project_name = DEFAULT_PROJECT_NAME
            st.session_state.editor_ver += 1
            st.session_state.last_view_index = None
            st.session_state.uploaded_token = None
            save_data()
            st.rerun()

        st.divider()
        if st.session_state.save_error:
            st.error(f"存檔失敗：{st.session_state.save_error}")
        elif st.session_state.last_saved:
            st.caption(f"✅ 已自動存檔 {st.session_state.last_saved}")
            st.caption(f"存放位置：{DATA_FILE}")
        else:
            st.caption(f"存放位置：{DATA_FILE}")

    # ----- Header + metrics -----
    df = st.session_state.df
    metrics = calc_metrics(df)

    st.title(APP_NAME)
    st.caption(
        f"專案：{st.session_state.project_name}"
        f"　｜　共 {len(df)} 筆任務（主任務 {metrics['主任務數']} / 子任務 {metrics['子任務數']}）"
        f"　｜　今天 {date.today():%Y-%m-%d}"
    )

    cols = st.columns(8)
    cols[0].metric("主任務數", metrics["主任務數"])
    cols[1].metric("已完成", metrics["已完成"])
    cols[2].metric("進行中", metrics["進行中"])
    cols[3].metric("延遲", metrics["延遲"])
    cols[4].metric("未開始", metrics["未開始"])
    cols[5].metric("完成率", f"{metrics['完成率']:.0f}%")
    cols[6].metric("整體進度", f"{metrics['整體進度']:.0f}%")
    cols[7].metric("已逾期", metrics["逾期"], help="截止日期已過且狀態不是「已完成」")

    st.divider()

    tab1, tab2, tab3 = st.tabs(["📝 任務清單", "📅 時程甘特圖", "📥 匯出"])

    # ===== Tab 1: Editable list =====
    with tab1:
        st.markdown("#### 任務清單（可直接編輯）")
        st.caption("支援新增列、刪除列；修改會自動存到本機檔案。有套用篩選 / 收合時，刪除的列也會真的從資料中移除。")

        # ---- 主任務收合控制 ----
        mains = collapsible_mains(df)
        labels = [label for _, label, _ in mains]
        label_to_wbs = {label: wbs for wbs, label, _ in mains}

        # 主任務清單有變動（改名 / 新增 / 刪除）時，換掉 widget key 避免殘留舊選項
        if labels != st.session_state.expand_options:
            st.session_state.expand_options = labels
            st.session_state.expand_ver += 1

        if mains:
            head, btn_a, btn_b = st.columns([6, 1, 1])
            head.caption("點選主任務可展開 / 收合其子任務（灰色＝已收合）")
            if btn_a.button("全部展開", width="stretch"):
                st.session_state.collapsed_wbs = set()
                st.session_state.expand_ver += 1
                st.rerun()
            if btn_b.button("全部收合", width="stretch"):
                st.session_state.collapsed_wbs = {wbs for wbs, _, _ in mains}
                st.session_state.expand_ver += 1
                st.rerun()

            selected = st.pills(
                "展開的主任務",
                options=labels,
                selection_mode="multi",
                default=[l for l in labels if label_to_wbs[l] not in st.session_state.collapsed_wbs],
                key=f"expand_pills_{st.session_state.expand_ver}",
                label_visibility="collapsed",
            )
            st.session_state.collapsed_wbs = {
                wbs for wbs, label, _ in mains if label not in (selected or [])
            }

        collapsed = st.session_state.collapsed_wbs

        # ---- 篩選 + 收合 → 實際顯示的列 ----
        mask = (
            df["層級"].isin(filter_level)
            & df["狀態"].isin(filter_status)
            & df["優先順序"].isin(filter_priority)
            & collapse_mask(df, collapsed)
        )
        view_index = df.index[mask].tolist()
        view_df = df.loc[view_index].reset_index(drop=True)
        # 顯示的列不是全部時，編輯位置會對不上主資料，套用編輯後必須重建 editor
        st.session_state.filter_active = len(view_index) != len(df)

        # 顯示的列組成一變（篩選 / 收合），舊的編輯狀態位置就失效，
        # 換 key 重建 editor，否則下一次會把編輯套用到錯的列
        if (
            st.session_state.last_view_index is not None
            and view_index != st.session_state.last_view_index
        ):
            st.session_state.editor_ver += 1

        view_df.insert(0, "階層", tree_markers(view_df, collapsed, child_counts(df)))

        column_config = {
            "階層": st.column_config.TextColumn("階層", width="small", disabled=True),
            "WBS編號": st.column_config.TextColumn("WBS編號", width="small"),
            "層級": st.column_config.SelectboxColumn("層級", options=LEVEL_OPTIONS, required=True, width="small"),
            "任務名稱": st.column_config.TextColumn("任務名稱", width="medium"),
            "負責人": st.column_config.TextColumn("負責人", width="small"),
            "開始日期": st.column_config.DateColumn("開始日期", format="YYYY-MM-DD", width="small"),
            "截止日期": st.column_config.DateColumn("截止日期", format="YYYY-MM-DD", width="small"),
            "狀態": st.column_config.SelectboxColumn("狀態", options=STATUS_OPTIONS, required=True, width="small"),
            "優先順序": st.column_config.SelectboxColumn("優先順序", options=PRIORITY_OPTIONS, required=True, width="small"),
            "進度(%)": st.column_config.NumberColumn("進度(%)", min_value=0, max_value=100, step=5, format="%d%%", width="small"),
            "備註": st.column_config.TextColumn("備註", width="large"),
        }

        st.data_editor(
            view_df,
            column_config=column_config,
            num_rows="dynamic",
            hide_index=True,
            width="stretch",
            key=editor_key(),
        )
        # 供下一輪把編輯位置對應回主資料
        st.session_state.last_view_index = view_index

        if len(view_index) != len(df):
            hidden_by_collapse = int((~collapse_mask(df, collapsed)).sum())
            note = f"目前只顯示 {len(view_index)} / {len(df)} 筆"
            if hidden_by_collapse:
                note += f"（其中 {hidden_by_collapse} 筆是被收合的子任務）"
            st.info(note + "。統計數字與甘特圖一律以全部資料計算。")

        orphans = orphan_subtasks(df)
        if len(orphans):
            st.caption(
                "⚠ 以下子任務找不到對應的主任務（主任務被刪掉時子任務不會跟著刪，避免誤刪），"
                "所以不會被收合：" + "、".join(
                    f"{r['WBS編號']} {r['任務名稱']}" for _, r in orphans.iterrows()
                )
            )

        overdue = df[overdue_mask(df)]
        if len(overdue):
            st.warning(
                "已逾期：" + "、".join(
                    f"{r['WBS編號']} {r['任務名稱']}（{pd.Timestamp(r['截止日期']):%m/%d}）"
                    for _, r in overdue.sort_values("截止日期").iterrows()
                )
            )

    # ===== Tab 2: Gantt =====
    with tab2:
        st.markdown("#### 時程甘特圖")
        st.caption("紅色虛線為今天。子任務與日期沿用設定可在左側調整。")

        fig = build_gantt(df, show_subtasks=show_sub_in_gantt, inherit_parent_dates=inherit_dates)
        if fig is None:
            st.warning("目前沒有同時具備開始與截止日期的任務，無法繪製甘特圖。請在任務清單補上日期。")
        else:
            st.plotly_chart(fig, width="stretch")
            legend_cols = st.columns(len(STATUS_COLORS))
            for i, (status, color) in enumerate(STATUS_COLORS.items()):
                legend_cols[i].markdown(
                    f"<div style='background:{color};color:white;padding:4px 8px;"
                    f"border-radius:4px;text-align:center;font-size:13px'>{status}</div>",
                    unsafe_allow_html=True,
                )

    # ===== Tab 3: Export =====
    with tab3:
        st.markdown("#### 匯出資料")
        stamp = datetime.now().strftime("%Y%m%d")
        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown("**Excel 檔（含格式，日期為真正的日期格式）**")
            st.download_button(
                label="⬇️ 下載 Excel",
                data=to_excel_bytes(df, st.session_state.project_name),
                file_name=f"{APP_NAME}_{st.session_state.project_name}_{stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

        with col_b:
            st.markdown("**CSV 檔（UTF-8 with BOM，Excel 開不會亂碼）**")
            csv_df = pd.DataFrame(df_to_records(df))
            st.download_button(
                label="⬇️ 下載 CSV",
                data=csv_df.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"{APP_NAME}_{st.session_state.project_name}_{stamp}.csv",
                mime="text/csv",
                width="stretch",
            )

        st.divider()
        st.markdown("#### 目前資料預覽")
        st.dataframe(df, width="stretch", hide_index=True)


if __name__ == "__main__":
    main()
